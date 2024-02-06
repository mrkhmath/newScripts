import pandas as pd
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Image, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from reportlab.lib import colors
import datetime
from openpyxl import load_workbook
from reportlab.lib.units import inch  # Importing inch for measurements

# Load the workbook and select the final sheet
wb = load_workbook('C:/Users/kh_ma/Downloads/reportsFile10.xlsx', data_only=True)
ws = wb['final']

# Define the path for the header image
header_image_path = 'C:/Users/kh_ma/Documents/madrasa/src/examHead.png'

# Prepare styles
styles = getSampleStyleSheet()
boldStyle = ParagraphStyle('boldStyle', parent=styles['Normal'], fontName='Helvetica-Bold')

# Create a PDF for each student
for row in ws.iter_rows(min_row=2, values_only=True):
    esis_number = row[0]
    student_name = row[1]
    student_class = row[2]
    yes_indices = [i for i, x in enumerate(row[3:], start=4) if x and str(x).strip().lower() == 'yes']

    # Create a document for each student
    pdf_path = f'C:/Users/kh_ma/Downloads/resitLetters10/{esis_number}.pdf'
    doc = SimpleDocTemplate(pdf_path, pagesize=letter)
    elements = []

    # Add the header image
    img = Image(header_image_path, width=468, height=60)
    img.hAlign = 'CENTER'
    elements.append(img)
    elements.append(Spacer(1, 12))

    # Add the title
    elements.append(Paragraph("Resit Exam Acknowledgment", styles['Title']))
    elements.append(Spacer(1, 24))

    # Add ESIS Number and Student Name in bold
    elements.append(Paragraph(f"<b>ESIS Number:</b> {esis_number}", boldStyle))
    elements.append(Spacer(1, 12))  # Double space
    elements.append(Paragraph(f"<b>Student Name:</b> {student_name}", boldStyle))
    elements.append(Spacer(1, 24))  # Double space

    # Additional paragraphs
    text_lines = [
        "Remember, this resit exam is an opportunity for growth and learning.",
        "Regardless of past performances, focus on giving your best effort this time.",
        "Our faculty and support staff are here to assist you,",
        "So please do not hesitate to reach out if you have any questions or need further assistance.",
        "We believe in your ability to succeed and look forward to seeing you excel in your resit exam.",
        "Wishing you all the best in your preparations and the exam itself.",
        "Sincerely,"
    ]
    for line in text_lines:
        elements.append(Paragraph(line, styles['Normal']))
        elements.append(Spacer(1, 12))

    # Create a table for the subjects and dates
    data = [['Subject', 'Date']]
    exam_date = datetime.date(2024, 2, 4)
    for index in yes_indices:
        subject = ws.cell(row=1, column=index).value
        exam_date += datetime.timedelta(days=1)  # Increment the day for each 'yes'
        data.append([subject,exam_date.strftime('%A') + " " + exam_date.strftime('%d/%m/%Y') ])

    # Add the table to the elements list
    t = Table(data, colWidths=[None, 2.5*inch], hAlign='LEFT')
    t.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige)
    ]))
    elements.append(t)

    # Build the PDF
    doc.build(elements)

# Close the workbook
wb.close()
