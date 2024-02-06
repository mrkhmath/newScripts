from openpyxl import load_workbook

# Load the workbook and select the active sheet
wb_path = 'C:/Users/kh_ma/Downloads/rist10sched.xlsx'
wb = load_workbook(wb_path)
sheet = wb.active  # Assuming the data is in the active sheet

# Create a new sheet for the modified data
if "Modified" not in wb.sheetnames:
    wb.create_sheet("Modified")
new_sheet = wb["Modified"]

# Get the headers (subjects) from row 1, starting from column D
subjects = [cell.value for cell in sheet[1][3:]]

# Copy headers to the new sheet
for col, header in enumerate(sheet[1], start=1):
    new_sheet.cell(row=1, column=col, value=header.value)

# Iterate over the rows, starting from the second row
for row_index, row in enumerate(sheet.iter_rows(min_row=2, max_col=sheet.max_column), start=2):
    # Preserve the first three columns as they are
    for col_index in range(1, 4):
        new_sheet.cell(row=row_index, column=col_index, value=row[col_index - 1].value)
    
    # Process remaining columns, starting from the fourth
    col_data = []  # List to hold non-empty data for shifting to the left
    for cell in row[3:]:  # Skip the first three columns
        if cell.value and str(cell.value).strip().lower() == 'yes':
            # Replace "yes" with the subject name from the header
            subject_name = subjects[cell.column - 4]  # Adjust for zero-based index
            col_data.append(subject_name)
        else:
            col_data.append(None)  # Placeholder for non-"yes" values
    
    # Shift non-empty (non-None) data to the left and write to the new sheet
    col_data = [item for item in col_data if item is not None]  # Remove None values
    for i, value in enumerate(col_data, start=4):  # Start writing from the fourth column
        new_sheet.cell(row=row_index, column=i, value=value)

# Save the workbook with modifications
wb.save('C:/Users/kh_ma/Downloads/rist10sched_modified.xlsx')
print("Workbook saved with modifications.")

# Clean up
wb.close()
